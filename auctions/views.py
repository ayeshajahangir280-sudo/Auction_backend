import csv
import json
import random
import re
import time
import uuid
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.conf import settings
from django.contrib.auth import authenticate, get_user_model
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db import transaction
from django.db.models import Count, F, Prefetch, Sum
from django.http import Http404, HttpResponse, StreamingHttpResponse
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import mixins, status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.utils.encoders import JSONEncoder
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken

from .models import Auction, AuctionLog, AuctionSettings, Bid, Category, Player, RoleProfile, SoldPlayer, Sponsor, Team, TeamCategoryLimit, TeamOwner, UploadedImage
from .permissions import is_auction_manager, is_super_admin, is_team_owner, scoped_auction_for_user
from .pdf import build_team_roster_pdf, team_roster_pdf_filename
from .serializers import (
    AuctionLogSerializer,
    AuctionSerializer,
    BidSerializer,
    CategorySerializer,
    PlayerSerializer,
    SoldPlayerSerializer,
    SponsorSerializer,
    TeamOwnerSerializer,
    TeamSerializer,
    UserSummarySerializer,
)

User = get_user_model()
LIVE_EVENT_POLL_SECONDS = 0.1

MISSING_IMAGE_SVG = b"""<svg xmlns="http://www.w3.org/2000/svg" width="320" height="240" viewBox="0 0 320 240"><rect width="320" height="240" fill="#e7edf8"/><path d="M70 158l48-56 42 48 26-30 64 74H70z" fill="#b8c7df"/><circle cx="226" cy="78" r="22" fill="#c9d6ea"/><text x="160" y="218" text-anchor="middle" font-family="Arial, sans-serif" font-size="18" font-weight="700" fill="#48618a">Image unavailable</text></svg>"""


def build_media_url(request, path: str) -> str:
    media_url = settings.MEDIA_URL if settings.MEDIA_URL.startswith("/") else f"/{settings.MEDIA_URL}"
    return request.build_absolute_uri(f"{media_url.rstrip('/')}/{path.lstrip('/')}")


def serve_uploaded_media(request, path: str):
    clean_path = path.lstrip("/")
    uploaded_image = UploadedImage.objects.filter(path=clean_path).first()
    if uploaded_image:
        response = HttpResponse(bytes(uploaded_image.data), content_type=uploaded_image.content_type)
        response["Cache-Control"] = "public, max-age=31536000, immutable"
        response["Content-Length"] = str(uploaded_image.size)
        return response

    if default_storage.exists(clean_path):
        with default_storage.open(clean_path, "rb") as stored_file:
            data = stored_file.read()
        content_type = "image/svg+xml" if clean_path.lower().endswith(".svg") else "image/jpeg"
        response = HttpResponse(data, content_type=content_type)
        response["Cache-Control"] = "public, max-age=86400"
        response["Content-Length"] = str(len(data))
        return response

    if clean_path.lower().startswith("uploads/images/"):
        response = HttpResponse(MISSING_IMAGE_SVG, content_type="image/svg+xml")
        response["Cache-Control"] = "no-store"
        return response

    raise Http404("Media file not found.")

PLAYER_IMPORT_HEADER_ALIASES = {
    "first": "first_name",
    "firstname": "first_name",
    "first_name": "first_name",
    "last": "last_name",
    "lastname": "last_name",
    "last_name": "last_name",
    "name": "full_name",
    "fullname": "full_name",
    "full_name": "full_name",
    "playername": "full_name",
    "player_name": "full_name",
    "category": "category",
    "categoryid": "category",
    "category_id": "category",
    "categorycode": "category",
    "category_code": "category",
    "categorynumber": "category",
    "category_number": "category",
    "categoryno": "category",
    "category_no": "category",
    "categoryname": "category",
    "category_name": "category",
    "catid": "category",
    "cat_id": "category",
    "catcode": "category",
    "cat_code": "category",
    "image": "image_url",
    "image_link": "image_url",
    "imagelink": "image_url",
    "imageurl": "image_url",
    "image_url": "image_url",
    "image_urls": "image_url",
    "img": "image_url",
    "imgurl": "image_url",
    "img_url": "image_url",
    "photo": "image_url",
    "photo_url": "image_url",
    "photourl": "image_url",
    "picture": "image_url",
    "picture_url": "image_url",
    "profile": "image_url",
    "profile_image": "image_url",
    "profile_photo": "image_url",
    "playerimage": "image_url",
    "player_image": "image_url",
    "player_photo": "image_url",
    "player_picture": "image_url",
    "fileurl": "image_url",
    "file_url": "image_url",
    "uploadurl": "image_url",
    "upload_url": "image_url",
    "attachmenturl": "image_url",
    "attachment_url": "image_url",
    "photo_file_url": "image_url",
    "role": "role",
    "type": "role",
    "playerrole": "role",
    "player_role": "role",
    "country": "country",
    "nation": "country",
    "age": "age",
    "baseprice": "base_price",
    "base_price": "base_price",
    "basevalue": "base_price",
    "base_value": "base_price",
    "price": "base_price",
    "order": "queue_order",
    "queueorder": "queue_order",
    "queue_order": "queue_order",
    "srno": "queue_order",
    "sr_no": "queue_order",
}

PLAYER_ROLE_ALIASES = {
    "bat": Player.Role.BATTER,
    "batter": Player.Role.BATTER,
    "bowler": Player.Role.BOWLER,
    "allrounder": Player.Role.ALL_ROUNDER,
    "all_rounder": Player.Role.ALL_ROUNDER,
    "wicketkeeper": Player.Role.WICKET_KEEPER,
    "wicket_keeper": Player.Role.WICKET_KEEPER,
    "keeper": Player.Role.WICKET_KEEPER,
}


def clean_import_cell(value) -> str:
    if hasattr(value, "value"):
        value = value.value
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_import_url(value) -> str:
    text = clean_import_cell(value)
    hyperlink = getattr(value, "hyperlink", None)
    if hyperlink and getattr(hyperlink, "target", None):
        text = clean_import_cell(hyperlink.target)
    if text.lower().startswith("www."):
        return f"https://{text}"
    return text


def normalize_import_key(value) -> str:
    return re.sub(r"[^a-z0-9]+", "_", clean_import_cell(value).lower()).strip("_")


def canonical_import_key(value) -> str | None:
    key = normalize_import_key(value)
    return PLAYER_IMPORT_HEADER_ALIASES.get(key) or PLAYER_IMPORT_HEADER_ALIASES.get(key.replace("_", ""))


def parse_import_decimal(value, default: str = "0") -> Decimal:
    text = clean_import_cell(value).replace(",", "")
    if not text:
        text = default
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError("Base price must be a number.") from exc


def parse_import_int(value, default: int = 0) -> int:
    text = clean_import_cell(value)
    if not text:
        return default
    try:
        return int(float(text))
    except ValueError as exc:
        raise ValueError("Age and order must be numbers.") from exc


def resolve_import_role(value) -> str:
    text = normalize_import_key(value)
    if not text:
        return ""
    role = PLAYER_ROLE_ALIASES.get(text) or PLAYER_ROLE_ALIASES.get(text.replace("_", ""))
    if not role:
        raise ValueError(f"Role '{clean_import_cell(value)}' is not valid.")
    return role


def normalize_category_display_name(value) -> str:
    return re.sub(r"\s+", " ", clean_import_cell(value)).strip()


def normalize_category_lookup_key(value) -> str:
    return normalize_category_display_name(value).casefold()


def resolve_import_category(auction: Auction, value, base_value: Decimal | None = None):
    text = clean_import_cell(value)
    category_name = normalize_category_display_name(text)
    if not category_name:
        category, created = Category.objects.get_or_create(
            auction=auction,
            name="Uncategorized",
            defaults={"base_value": Decimal("0"), "color": "#64748B"},
        )
        return category, created

    category = auction.categories.filter(category_id__iexact=text).first()
    if not category and text.isdigit():
        category = auction.categories.filter(pk=int(text)).first()

    lookup_key = normalize_category_lookup_key(category_name)
    if not category:
        for existing_category in auction.categories.all():
            if normalize_category_lookup_key(existing_category.name) == lookup_key:
                category = existing_category
                break

    if category:
        if base_value and base_value > 0 and category.base_value == 0:
            category.base_value = base_value
            category.save(update_fields=["base_value"])
        return category, False

    category = Category.objects.create(
        auction=auction,
        name=category_name[:120],
        base_value=base_value or Decimal("0"),
    )
    return category, True


def resolve_auction_category(auction: Auction, value, field_name: str) -> Category:
    text = clean_import_cell(value)
    if not text:
        raise ValidationError({field_name: "Choose a category."})

    category = auction.categories.filter(category_id__iexact=text).first()
    if not category and text.isdigit():
        category = auction.categories.filter(pk=int(text)).first()
    if not category:
        raise ValidationError({field_name: "Category not found for this auction."})
    return category


def user_can_access_auction(user, auction: Auction) -> bool:
    if is_super_admin(user):
        return True
    scoped = scoped_auction_for_user(user)
    return bool(scoped and scoped.pk == auction.pk)


def require_auction_staff(user) -> None:
    if not (is_super_admin(user) or is_auction_manager(user)):
        raise PermissionDenied("Only Super Admin or the assigned auction manager can perform this action.")


def prevent_live_cache(response):
    response["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"
    return response


def live_response(data, status_code=status.HTTP_200_OK):
    return prevent_live_cache(Response(data, status=status_code))


def bump_live_revision(auction: Auction) -> None:
    Auction.objects.filter(pk=auction.pk).update(live_revision=F("live_revision") + 1, updated_at=timezone.now())
    auction.refresh_from_db(fields=["live_revision", "updated_at"])


def sse_event(event: str, data: dict, event_id: str | int | None = None) -> str:
    lines = []
    if event_id is not None:
        lines.append(f"id: {event_id}")
    lines.append(f"event: {event}")
    lines.append(f"data: {json.dumps(data, separators=(',', ':'), cls=JSONEncoder)}")
    return "\n".join(lines) + "\n\n"


def sse_response(stream):
    response = StreamingHttpResponse(stream, content_type="text/event-stream")
    prevent_live_cache(response)
    response["X-Accel-Buffering"] = "no"
    return response


def highest_approved_bid(auction: Auction, player: Player | None = None):
    qs = auction.bids.select_related("player", "player__category", "team").filter(bid_status=Bid.Status.APPROVED)
    if player:
        qs = qs.filter(player=player)
    return qs.order_by("-bid_amount", "-approved_at", "-created_at").first()


def highest_active_bid(auction: Auction, player: Player | None = None):
    qs = auction.bids.select_related("player", "player__category", "team").filter(bid_status__in=[Bid.Status.PENDING, Bid.Status.APPROVED])
    if player:
        qs = qs.filter(player=player)
    return qs.order_by("-bid_amount", "-created_at").first()


def highest_active_bid_for_team(auction: Auction, player: Player, team: Team):
    return (
        auction.bids.select_related("player", "player__category", "team")
        .filter(
            player=player,
            team=team,
            bid_status__in=[Bid.Status.PENDING, Bid.Status.APPROVED],
        )
        .order_by("-bid_amount", "-created_at")
        .first()
    )


def remaining_required_points(team: Team) -> Decimal:
    total = Decimal("0")
    cache = getattr(team, "_prefetched_objects_cache", {})
    limits = cache.get("category_limits")
    if limits is None:
        limits = team.category_limits.select_related("category")
    for limit in limits:
        prefetched_players = cache.get("players")
        if prefetched_players is None:
            bought_count = team.players.filter(category=limit.category).count()
        else:
            bought_count = sum(1 for player in prefetched_players if player.category_id == limit.category_id)
        total += limit.category.base_value * max(limit.maximum_players - bought_count, 0)
    return total


def available_bid_budget(team: Team) -> Decimal:
    return team.remaining_purse - remaining_required_points(team)


def team_category_limit_message(team: Team, player: Player) -> str:
    if not player.category_id:
        return ""
    cache = getattr(team, "_prefetched_objects_cache", {})
    prefetched_limits = cache.get("category_limits")
    if prefetched_limits is not None:
        limit = next((item for item in prefetched_limits if item.category_id == player.category_id), None)
    else:
        limit = team.category_limits.filter(category=player.category).first()
    if not limit or limit.maximum_players == 0:
        return ""
    prefetched_players = cache.get("players")
    bought_count = (
        sum(1 for item in prefetched_players if item.category_id == player.category_id)
        if prefetched_players is not None
        else team.players.filter(category=player.category).count()
    )
    if bought_count >= limit.maximum_players:
        return (
            f"Category limit reached: {team.short_name} already has the maximum players "
            f"for {player.category.name} ({bought_count}/{limit.maximum_players})."
        )
    return ""


def team_player_limit_message(team: Team, player: Player) -> tuple[str, str]:
    prefetched_players = getattr(team, "_prefetched_objects_cache", {}).get("players")
    players_bought = len(prefetched_players) if prefetched_players is not None else team.players.count()
    if team.maximum_players and players_bought >= team.maximum_players:
        return "team", f"{team.short_name} already has the maximum squad of {team.maximum_players} players."
    category_message = team_category_limit_message(team, player)
    if category_message:
        return "category", category_message
    return "", ""


def validate_team_category_limit(team: Team, player: Player) -> None:
    message = team_category_limit_message(team, player)
    if message:
        raise ValidationError({"category": message})


def validate_team_player_limits(team: Team, player: Player) -> None:
    field, message = team_player_limit_message(team, player)
    if message:
        raise ValidationError({field: message})


def sell_current_player_to_team(
    auction: Auction,
    player: Player,
    team: Team,
    sold_price: Decimal,
    actor=None,
    winning_bid: Bid | None = None,
) -> None:
    if player.status == Player.Status.SOLD:
        return
    validate_team_player_limits(team, player)
    budget = available_bid_budget(team)
    if sold_price > budget:
        raise ValidationError(
            {"sold_price": f"Winning team must keep reserved purse. Available bid budget is {budget}."}
        )

    player.status = Player.Status.SOLD
    player.sold_team = team
    player.sold_price = sold_price
    player.save(update_fields=["status", "sold_team", "sold_price"])
    team.remaining_purse -= sold_price
    team.players_bought += 1
    team.save(update_fields=["remaining_purse", "players_bought"])
    SoldPlayer.objects.update_or_create(
        auction=auction,
        player=player,
        defaults={"team": team, "sold_price": sold_price},
    )
    pending_bids = Bid.objects.filter(auction=auction, player=player, bid_status=Bid.Status.PENDING)
    if winning_bid:
        pending_bids = pending_bids.exclude(pk=winning_bid.pk)
    pending_bids.update(bid_status=Bid.Status.REJECTED, approved_by_admin=actor, approved_at=timezone.now())
    if auction.current_player_id == player.pk:
        auction.current_player = None
        auction.sold_animation_state = False
        auction.save(update_fields=["current_player", "sold_animation_state"])
    bump_live_revision(auction)
    AuctionLog.objects.create(auction=auction, actor=actor, action="player.sold", message=f"{player.full_name} sold to {team.short_name}.")


def live_team_queryset(auction: Auction):
    return auction.teams.select_related("auction", "owner_user").prefetch_related(
        Prefetch("category_limits", queryset=TeamCategoryLimit.objects.select_related("category")),
        Prefetch("players", queryset=Player.objects.only("id", "sold_team_id", "category_id")),
    )


def live_bid_queryset(auction: Auction):
    return auction.bids.select_related("player", "player__category", "team").prefetch_related(
        Prefetch(
            "team__category_limits",
            queryset=TeamCategoryLimit.objects.select_related("category"),
        ),
        Prefetch("team__players", queryset=Player.objects.only("id", "sold_team_id", "category_id")),
    )


def live_auction_queryset():
    return Auction.objects.select_related(
        "current_player",
        "current_player__category",
        "settings",
    ).prefetch_related("sponsors")


def live_auction_for_pk(auction_pk: int):
    return live_auction_queryset().filter(pk=auction_pk).first()


def decimal_text(value) -> str:
    return str(value if value is not None else Decimal("0"))


def iso_datetime(value):
    if value is None:
        return None
    return value.isoformat().replace("+00:00", "Z")


def live_settings_data(auction: Auction) -> dict:
    try:
        auction_settings = auction.settings
    except AuctionSettings.DoesNotExist:
        return {
            "show_remaining_purse": True,
            "require_bid_approval": True,
            "enable_owner_bidding": False,
            "auto_advance_after_sale": False,
            "sponsor_rotation_seconds": 8,
            "public_screen_theme": "purple_glow",
            "notes": "",
        }
    return {
        "show_remaining_purse": auction_settings.show_remaining_purse,
        "require_bid_approval": auction_settings.require_bid_approval,
        "enable_owner_bidding": auction_settings.enable_owner_bidding,
        "auto_advance_after_sale": auction_settings.auto_advance_after_sale,
        "sponsor_rotation_seconds": auction_settings.sponsor_rotation_seconds,
        "public_screen_theme": auction_settings.public_screen_theme,
        "notes": auction_settings.notes,
    }


def live_sponsor_data(sponsor: Sponsor) -> dict:
    return {
        "id": sponsor.pk,
        "auction": sponsor.auction_id,
        "name": sponsor.name,
        "logo_url": sponsor.logo_url,
        "status": sponsor.status,
        "sort_order": sponsor.sort_order,
    }


def live_auction_data(auction: Auction, results: dict, team_count: int) -> dict:
    return {
        "id": auction.pk,
        "auction_id": auction.auction_id,
        "name": auction.name,
        "manager": auction.manager_id,
        "manager_name": "",
        "auction_type": auction.auction_type,
        "number_of_teams": auction.number_of_teams,
        "allotted_to_user": auction.allotted_to_user,
        "payment_status": auction.payment_status,
        "status": auction.status,
        "logo_url": auction.logo_url,
        "unit": auction.unit,
        "purse_amount": decimal_text(auction.purse_amount),
        "purse": decimal_text(auction.purse_amount),
        "purse_type": auction.unit,
        "bid_increment": decimal_text(auction.bid_increment),
        "timer_duration": auction.timer_duration,
        "minimum_players_per_team": auction.minimum_players_per_team,
        "maximum_players_per_team": auction.maximum_players_per_team,
        "current_player": auction.current_player_id,
        "sold_animation_state": auction.sold_animation_state,
        "live_revision": auction.live_revision,
        "team_count": team_count,
        "player_count": results["sold_count"] + results["unsold_count"] + results["available_count"],
        "setup_enabled": True,
        "sponsors": [live_sponsor_data(sponsor) for sponsor in auction.sponsors.all()],
        "settings": live_settings_data(auction),
        "created_at": iso_datetime(auction.created_at),
        "updated_at": iso_datetime(auction.updated_at),
    }


def live_player_data(player: Player | None) -> dict | None:
    if not player:
        return None
    return {
        "id": player.pk,
        "auction": player.auction_id,
        "player_id": player.player_id,
        "first_name": player.first_name,
        "last_name": player.last_name,
        "full_name": player.full_name,
        "category": player.category_id,
        "category_name": player.category.name if player.category else "",
        "image_url": player.image_url,
        "role": player.role,
        "country": player.country,
        "age": player.age,
        "base_price": decimal_text(player.base_price),
        "extra_field_1": player.extra_field_1,
        "extra_field_2": player.extra_field_2,
        "extra_field_3": player.extra_field_3,
        "extra_field_4": player.extra_field_4,
        "status": player.status,
        "sold_team": player.sold_team_id,
        "sold_team_name": getattr(player.sold_team, "name", "") if player.sold_team_id else "",
        "sold_price": decimal_text(player.sold_price) if player.sold_price is not None else None,
        "queue_order": player.queue_order,
    }


def live_team_limit_data(limit: TeamCategoryLimit) -> dict:
    team = limit.team
    prefetched_players = getattr(team, "_prefetched_objects_cache", {}).get("players")
    bought_count = (
        sum(1 for player in prefetched_players if player.category_id == limit.category_id)
        if prefetched_players is not None
        else team.players.filter(category=limit.category).count()
    )
    remaining_slots = max(limit.maximum_players - bought_count, 0)
    required_points = limit.category.base_value * remaining_slots
    return {
        "id": limit.pk,
        "team": team.pk,
        "category": limit.category_id,
        "category_name": limit.category.name,
        "category_base_value": decimal_text(limit.category.base_value),
        "maximum_players": limit.maximum_players,
        "bought_count": bought_count,
        "remaining_slots": remaining_slots,
        "required_points": decimal_text(required_points),
    }


def live_team_data(team: Team) -> dict:
    required = remaining_required_points(team)
    return {
        "id": team.pk,
        "auction": team.auction_id,
        "team_id": team.team_id,
        "name": team.name,
        "short_name": team.short_name,
        "logo_url": team.logo_url,
        "owner_name": team.owner_name,
        "owner_username": team.owner_username,
        "owner_user_id": team.owner_user_id,
        "purse_amount": decimal_text(team.auction.purse_amount),
        "purse_type": team.auction.unit,
        "remaining_purse": decimal_text(team.remaining_purse),
        "players_bought": team.players_bought,
        "maximum_players": team.maximum_players,
        "category_limits": [live_team_limit_data(limit) for limit in team.category_limits.all()],
        "required_points": decimal_text(required),
        "budget_left_after_required": decimal_text(team.remaining_purse - required),
        "status": team.status,
    }


def live_bid_data(bid: Bid | None) -> dict | None:
    if not bid:
        return None
    _field, limit_message = team_player_limit_message(bid.team, bid.player)
    return {
        "id": bid.pk,
        "auction": bid.auction_id,
        "player": bid.player_id,
        "player_name": bid.player.full_name,
        "player_image_url": bid.player.image_url,
        "player_category": bid.player.category.name if bid.player.category else "",
        "team": bid.team_id,
        "team_name": bid.team.name,
        "team_short_name": bid.team.short_name,
        "team_logo_url": bid.team.logo_url,
        "owner_name": bid.team.owner_name,
        "bid_amount": decimal_text(bid.bid_amount),
        "bid_type": bid.bid_type,
        "bid_status": bid.bid_status,
        "created_at": iso_datetime(bid.created_at),
        "approved_by_admin": bid.approved_by_admin_id,
        "approved_at": iso_datetime(bid.approved_at),
        "team_limit_reached": bool(limit_message),
        "team_limit_message": limit_message,
    }


def live_sold_player_data(sold: SoldPlayer | None) -> dict | None:
    if not sold:
        return None
    return {
        "id": sold.pk,
        "auction": sold.auction_id,
        "player": sold.player_id,
        "player_name": sold.player.full_name,
        "player_image_url": sold.player.image_url,
        "player_role": sold.player.role,
        "player_category": sold.player.category.name if sold.player.category else "",
        "team": sold.team_id,
        "team_name": sold.team.name,
        "team_short_name": sold.team.short_name,
        "team_logo_url": sold.team.logo_url,
        "sold_price": decimal_text(sold.sold_price),
        "sold_time": iso_datetime(sold.sold_time),
    }


def build_results(auction: Auction, teams=None) -> dict:
    sold_qs = auction.sold_players.select_related("player", "player__category", "team")
    sold_summary = sold_qs.aggregate(count=Count("id"), total=Sum("sold_price"))
    sold_count = sold_summary["count"] or 0
    total_spend = sold_summary["total"] or Decimal("0")
    top_sale = sold_qs.order_by("-sold_price").first()
    player_status_counts = {
        item["status"]: item["count"]
        for item in auction.players.values("status").annotate(count=Count("id"))
    }
    result_teams = []
    team_iterable = teams if teams is not None else auction.teams.annotate(roster_count=Count("players")).order_by("name")
    for team in team_iterable:
        required = remaining_required_points(team)
        result_teams.append(
            {
                "team_id": team.team_id,
                "name": team.name,
                "short_name": team.short_name,
                "remaining_purse": team.remaining_purse,
                "players_bought": team.players_bought,
                "maximum_players": team.maximum_players,
                "required_points": required,
                "budget_left_after_required": team.remaining_purse - required,
            }
        )
    return {
        "sold_count": sold_count,
        "unsold_count": player_status_counts.get(Player.Status.UNSOLD, 0),
        "available_count": player_status_counts.get(Player.Status.AVAILABLE, 0),
        "total_spend": decimal_text(total_spend),
        "top_sale": live_sold_player_data(top_sale) if top_sale else None,
        "teams": result_teams,
    }


def serialize_live_state(auction: Auction, team_scope: Team | None = None) -> dict:
    current_player = auction.current_player
    bids_qs = live_bid_queryset(auction)
    current_player_bids = (
        bids_qs.filter(player=current_player, bid_status__in=[Bid.Status.PENDING, Bid.Status.APPROVED])
        .order_by("-bid_amount", "-created_at")
        if current_player
        else Bid.objects.none()
    )
    bid_feed = (
        bids_qs.filter(bid_status=Bid.Status.APPROVED)
        .order_by("-created_at")[:20]
    )
    pending_qs = (
        bids_qs.filter(bid_status=Bid.Status.PENDING)
        .order_by("-bid_amount", "-created_at")
    )
    teams_qs = live_team_queryset(auction).order_by("name")
    sold_players_qs = auction.sold_players.select_related("player", "player__category", "team")
    if team_scope:
        pending_qs = pending_qs.filter(team=team_scope)
        teams_qs = teams_qs.filter(pk=team_scope.pk)
        sold_players_qs = sold_players_qs.filter(team=team_scope)
    teams = list(teams_qs)
    results = build_results(auction, teams=teams)
    if team_scope:
        results["teams"] = [team for team in results["teams"] if team["team_id"] == team_scope.team_id]
        if results["top_sale"] and results["top_sale"]["team"] != team_scope.pk:
            results["top_sale"] = None
    pending = list(pending_qs[:20])
    current_player_bid_list = list(current_player_bids)
    bid_feed_list = list(bid_feed)
    sold_players = list(sold_players_qs[:10])
    current_bid = current_player_bid_list[0] if current_player_bid_list else None
    return {
        "auction": live_auction_data(auction, results, len(teams)),
        "current_player": live_player_data(current_player),
        "teams": [live_team_data(team) for team in teams],
        "current_bid": live_bid_data(current_bid),
        "current_player_bids": [live_bid_data(bid) for bid in current_player_bid_list],
        "pending_bids": [live_bid_data(bid) for bid in pending],
        "bid_feed": [live_bid_data(bid) for bid in bid_feed_list],
        "sold_players": [live_sold_player_data(sold) for sold in sold_players],
        "results": results,
    }


def close_current_player_before_switch(auction: Auction, next_player: Player, actor=None) -> None:
    current_player = auction.current_player
    if (
        not current_player
        or current_player.pk == next_player.pk
        or current_player.status != Player.Status.IN_AUCTION
    ):
        return

    has_active_bid = auction.bids.filter(
        player=current_player,
        bid_status__in=[Bid.Status.PENDING, Bid.Status.APPROVED],
    ).exists()
    if has_active_bid:
        current_player.status = Player.Status.AVAILABLE
        current_player.save(update_fields=["status"])
        return

    current_player.status = Player.Status.UNSOLD
    current_player.sold_team = None
    current_player.sold_price = None
    current_player.save(update_fields=["status", "sold_team", "sold_price"])
    AuctionLog.objects.create(
        auction=auction,
        actor=actor,
        action="player.auto_unsold",
        message=f"{current_player.full_name} marked unsold when the next player was selected.",
    )


def advance_to_random_player(auction: Auction, actor=None) -> Player | None:
    current_player_id = auction.current_player_id
    next_player_qs = auction.players.filter(status=Player.Status.AVAILABLE)
    if current_player_id:
        next_player_qs = next_player_qs.exclude(pk=current_player_id)
    available_count = next_player_qs.count()
    next_player = (
        next_player_qs.order_by("id")[random.randrange(available_count)]
        if available_count
        else None
    )
    if not next_player:
        current_player = auction.current_player
        if current_player and current_player.status == Player.Status.IN_AUCTION:
            return current_player
        auction.players.filter(status=Player.Status.IN_AUCTION).update(status=Player.Status.AVAILABLE)
        auction.current_player = None
        auction.sold_animation_state = False
        auction.save(update_fields=["current_player", "sold_animation_state"])
        return None

    previous_current_id = auction.current_player_id
    close_current_player_before_switch(auction, next_player, actor=actor)
    stale_in_auction_qs = auction.players.filter(status=Player.Status.IN_AUCTION).exclude(
        pk=next_player.pk
    )
    if previous_current_id:
        stale_in_auction_qs = stale_in_auction_qs.exclude(pk=previous_current_id)
    stale_in_auction_qs.update(status=Player.Status.AVAILABLE)
    next_player.status = Player.Status.IN_AUCTION
    next_player.save(update_fields=["status"])
    auction.current_player = next_player
    auction.status = Auction.Status.LIVE
    auction.sold_animation_state = False
    auction.save(update_fields=["current_player", "status", "sold_animation_state"])
    AuctionLog.objects.create(
        auction=auction,
        actor=actor,
        action="auction.next_player",
        message=f"{next_player.full_name} moved to the block randomly.",
    )
    return next_player


def public_active_auction():
    auction = (
        live_auction_queryset()
        .filter(status__in=[Auction.Status.LIVE, Auction.Status.ACTIVE, Auction.Status.SETUP])
        .order_by("-updated_at", "-created_at")
        .first()
    )
    if auction:
        return auction
    return (
        live_auction_queryset()
        .exclude(status=Auction.Status.ARCHIVED)
        .order_by("-updated_at", "-created_at")
        .first()
    )


def public_active_auction_token():
    token = (
        Auction.objects.filter(status__in=[Auction.Status.LIVE, Auction.Status.ACTIVE, Auction.Status.SETUP])
        .order_by("-updated_at", "-created_at")
        .values_list("pk", "auction_id", "live_revision")
        .first()
    )
    if token:
        return token
    return (
        Auction.objects.exclude(status=Auction.Status.ARCHIVED)
        .order_by("-updated_at", "-created_at")
        .values_list("pk", "auction_id", "live_revision")
        .first()
    )


def auction_revision_stream(auction_pk: int):
    auction = live_auction_for_pk(auction_pk)
    last_revision = auction.live_revision if auction else None
    yield "retry: 1500\n\n"
    if auction is not None:
        yield sse_event(
            "revision",
            {"revision": last_revision, "state": serialize_live_state(auction)},
            last_revision,
        )
    last_heartbeat = time.monotonic()
    while True:
        revision = Auction.objects.filter(pk=auction_pk).values_list("live_revision", flat=True).first()
        if revision is None:
            yield sse_event("deleted", {"auction": None})
            return
        if revision != last_revision:
            last_revision = revision
            auction = live_auction_for_pk(auction_pk)
            if auction is None:
                yield sse_event("deleted", {"auction": None})
                return
            yield sse_event(
                "revision",
                {"revision": revision, "state": serialize_live_state(auction)},
                revision,
            )
            last_heartbeat = time.monotonic()
        elif time.monotonic() - last_heartbeat >= 15:
            yield ": keepalive\n\n"
            last_heartbeat = time.monotonic()
        time.sleep(LIVE_EVENT_POLL_SECONDS)


def public_active_revision_stream():
    active_token = public_active_auction_token()
    auction = live_auction_for_pk(active_token[0]) if active_token else None
    last_token = f"{active_token[1]}:{active_token[2]}" if active_token else ""
    yield "retry: 1500\n\n"
    if auction:
        yield sse_event(
            "revision",
            {
                "auction_id": auction.auction_id,
                "revision": auction.live_revision,
                "state": serialize_live_state(auction),
            },
            last_token,
        )
    last_heartbeat = time.monotonic()
    while True:
        active_token = public_active_auction_token()
        token = f"{active_token[1]}:{active_token[2]}" if active_token else ""
        if token != last_token:
            last_token = token
            auction = live_auction_for_pk(active_token[0]) if active_token else None
            payload = (
                {
                    "auction_id": auction.auction_id,
                    "revision": auction.live_revision,
                    "state": serialize_live_state(auction),
                }
                if auction
                else {"auction_id": None}
            )
            yield sse_event("revision", payload, token or "none")
            last_heartbeat = time.monotonic()
        elif time.monotonic() - last_heartbeat >= 15:
            yield ": keepalive\n\n"
            last_heartbeat = time.monotonic()
        time.sleep(LIVE_EVENT_POLL_SECONDS)


class LoginView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        username = request.data.get("username", "").strip()
        password = request.data.get("password", "")
        user = authenticate(username=username, password=password)
        if not user:
            return Response({"detail": "Invalid username or password."}, status=status.HTTP_401_UNAUTHORIZED)
        refresh = RefreshToken.for_user(user)
        return Response(
            {
                "access": str(refresh.access_token),
                "refresh": str(refresh),
                "user": UserSummarySerializer(user).data,
            }
        )


class CurrentUserView(APIView):
    def get(self, request):
        return Response(UserSummarySerializer(request.user).data)


class ImageUploadView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            raise ValidationError({"file": "Choose an image file to upload."})

        content_type = getattr(uploaded_file, "content_type", "")
        if not content_type.startswith("image/"):
            raise ValidationError({"file": "Only image files can be uploaded."})

        suffix = Path(uploaded_file.name).suffix.lower()
        if suffix not in {".jpg", ".jpeg", ".png", ".gif", ".webp", ".avif", ".svg"}:
            suffix = ".jpg"
        file_data = uploaded_file.read()
        saved_path = f"uploads/images/{uuid.uuid4().hex}{suffix}"
        default_storage.save(saved_path, ContentFile(file_data))
        UploadedImage.objects.create(
            path=saved_path,
            content_type=content_type,
            data=file_data,
            size=len(file_data),
        )
        absolute_url = build_media_url(request, saved_path)
        return Response({"url": absolute_url, "path": saved_path}, status=status.HTTP_201_CREATED)


class UserViewSet(mixins.ListModelMixin, mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    serializer_class = UserSummarySerializer

    def get_queryset(self):
        qs = User.objects.select_related("role_profile", "role_profile__assigned_auction", "role_profile__team").order_by("username")
        if is_super_admin(self.request.user):
            role = self.request.query_params.get("role")
            if role:
                qs = qs.filter(role_profile__role=role)
            return qs
        return qs.filter(pk=self.request.user.pk)


class ScopedModelViewSet(viewsets.ModelViewSet):
    auction_field = "auction"

    def get_queryset(self):
        qs = super().get_queryset()
        if is_super_admin(self.request.user):
            auction_id = self.request.query_params.get("auction")
            return qs.filter(auction__auction_id=auction_id) if auction_id else qs
        if is_team_owner(self.request.user):
            return qs.none()
        auction = scoped_auction_for_user(self.request.user)
        if not auction:
            return qs.none()
        return qs.filter(auction=auction)

    def perform_create(self, serializer):
        require_auction_staff(self.request.user)
        auction = serializer.validated_data.get("auction") or scoped_auction_for_user(self.request.user)
        if not auction:
            raise PermissionDenied("Choose an auction or log in as an assigned auction user.")
        if not user_can_access_auction(self.request.user, auction):
            raise PermissionDenied("You cannot write data for this auction.")
        serializer.save(auction=auction)

    def perform_update(self, serializer):
        require_auction_staff(self.request.user)
        auction = serializer.validated_data.get("auction") or getattr(serializer.instance, "auction", None)
        if auction and not user_can_access_auction(self.request.user, auction):
            raise PermissionDenied("You cannot write data for this auction.")
        serializer.save()

    def perform_destroy(self, instance):
        require_auction_staff(self.request.user)
        auction = getattr(instance, "auction", None)
        if auction and not user_can_access_auction(self.request.user, auction):
            raise PermissionDenied("You cannot delete data for this auction.")
        instance.delete()


class AuctionViewSet(viewsets.ModelViewSet):
    queryset = Auction.objects.select_related(
        "manager",
        "current_player",
        "current_player__category",
        "settings",
    ).prefetch_related("sponsors")
    serializer_class = AuctionSerializer
    lookup_field = "auction_id"

    def get_permissions(self):
        public_actions = {"public_active", "public_live", "projector", "public_live_events", "public_active_events"}
        if self.action in public_actions:
            return [AllowAny()]
        return [IsAuthenticated()]

    def get_queryset(self):
        qs = super().get_queryset().annotate(
            team_total=Count("teams", distinct=True),
            player_total=Count("players", distinct=True),
        )
        if self.action in {"public_active", "public_live", "projector", "public_live_events", "public_active_events"}:
            return qs
        if is_super_admin(self.request.user):
            return qs
        if is_team_owner(self.request.user) and self.action not in {"live_state", "team_owner_bid", "team_roster", "team_roster_pdf"}:
            return qs.none()
        auction = scoped_auction_for_user(self.request.user)
        return qs.filter(pk=auction.pk) if auction else qs.none()

    def perform_create(self, serializer):
        if not is_super_admin(self.request.user):
            raise PermissionDenied("Only Super Admin can create auctions.")
        auction = serializer.save()
        AuctionLog.objects.create(
            auction=auction,
            actor=self.request.user,
            action="auction.created",
            message="Auction project was created.",
        )

    def perform_update(self, serializer):
        require_auction_staff(self.request.user)
        if not user_can_access_auction(self.request.user, serializer.instance):
            raise PermissionDenied("You cannot update this auction.")
        serializer.save()

    def perform_destroy(self, instance):
        require_auction_staff(self.request.user)
        if not user_can_access_auction(self.request.user, instance):
            raise PermissionDenied("You cannot delete this auction.")
        instance.delete()

    @action(detail=False, methods=["get"], url_path="manager-dashboard")
    def manager_dashboard(self, request):
        auction = scoped_auction_for_user(request.user)
        if not auction and is_super_admin(request.user):
            auction = self.get_queryset().first()
        if not auction:
            return Response({"detail": "No assigned auction found."}, status=status.HTTP_404_NOT_FOUND)
        teams = list(live_team_queryset(auction).order_by("name"))
        player_status_counts = {
            item["status"]: item["count"]
            for item in auction.players.values("status").annotate(count=Count("id"))
        }
        current_player = auction.current_player
        bid_qs = live_bid_queryset(auction)
        current_bid = highest_active_bid(auction, current_player) if current_player else None
        bid_feed = bid_qs.filter(bid_status=Bid.Status.APPROVED).order_by("-created_at")[:6]
        return Response(
            {
                "auction": AuctionSerializer(auction).data,
                "project_count": Auction.objects.count() if is_super_admin(request.user) else 1,
                "category_count": auction.categories.count(),
                "player_count": sum(player_status_counts.values()),
                "team_count": len(teams),
                "pending_bid_count": auction.bids.filter(bid_status=Bid.Status.PENDING).count(),
                "sold_count": player_status_counts.get(Player.Status.SOLD, 0),
                "results": build_results(auction, teams=teams),
                "current_player": live_player_data(current_player),
                "current_bid": live_bid_data(current_bid),
                "bid_feed": [live_bid_data(bid) for bid in bid_feed],
            }
        )

    @action(detail=True, methods=["get"], url_path="live-state")
    def live_state(self, request, auction_id=None):
        auction = self.get_object()
        team_scope = request.user.role_profile.team if is_team_owner(request.user) else None
        return live_response(serialize_live_state(auction, team_scope=team_scope))

    @action(detail=True, methods=["get"], url_path="public-live-events")
    def public_live_events(self, request, auction_id=None):
        auction = self.get_object()
        return sse_response(auction_revision_stream(auction.pk))

    @action(detail=False, methods=["get"], url_path="public-active-events")
    def public_active_events(self, request):
        return sse_response(public_active_revision_stream())

    @action(detail=True, methods=["get"], url_path="current-player")
    def current_player(self, request, auction_id=None):
        auction = self.get_object()
        return Response(PlayerSerializer(auction.current_player).data if auction.current_player else None)

    @action(detail=True, methods=["post"], url_path="start-auction")
    def start_auction(self, request, auction_id=None):
        require_auction_staff(request.user)
        auction = self.get_object()
        current_player_is_live = auction.current_player and auction.current_player.status == Player.Status.IN_AUCTION
        if not auction.players.filter(status=Player.Status.AVAILABLE).exists() and not current_player_is_live:
            raise ValidationError({"players": "Add available players before starting the auction."})
        if current_player_is_live:
            auction.status = Auction.Status.LIVE
            auction.save(update_fields=["status"])
        elif not advance_to_random_player(auction, request.user):
            raise ValidationError({"players": "No available players left to start."})
        AuctionLog.objects.create(
            auction=auction,
            actor=request.user,
            action="auction.started",
            message="Auction was started.",
        )
        bump_live_revision(auction)
        return live_response(serialize_live_state(auction))

    @transaction.atomic
    @action(detail=True, methods=["post"], url_path="complete-auction")
    def complete_auction(self, request, auction_id=None):
        require_auction_staff(request.user)
        auction = self.get_object()
        auction.players.filter(status__in=[Player.Status.AVAILABLE, Player.Status.IN_AUCTION]).update(
            status=Player.Status.UNSOLD
        )
        auction.current_player = None
        auction.status = Auction.Status.COMPLETED
        auction.sold_animation_state = False
        auction.save(update_fields=["current_player", "status", "sold_animation_state"])
        AuctionLog.objects.create(
            auction=auction,
            actor=request.user,
            action="auction.completed",
            message="Auction was marked completed.",
        )
        bump_live_revision(auction)
        return live_response(serialize_live_state(auction))

    @action(detail=True, methods=["post"], url_path="set-current-player")
    def set_current_player(self, request, auction_id=None):
        require_auction_staff(request.user)
        auction = self.get_object()
        player_code = request.data.get("player_id") or request.data.get("player")
        player = auction.players.filter(player_id=player_code).first() or auction.players.filter(pk=player_code).first()
        if not player:
            raise ValidationError({"player_id": "Player not found in this auction."})
        if player.status in {Player.Status.SOLD, Player.Status.UNSOLD}:
            raise ValidationError({"player_id": "Choose a player who is not sold or unsold."})
        previous_current_id = auction.current_player_id
        close_current_player_before_switch(auction, player, actor=request.user)
        stale_in_auction_qs = auction.players.filter(status=Player.Status.IN_AUCTION).exclude(
            pk=player.pk
        )
        if previous_current_id:
            stale_in_auction_qs = stale_in_auction_qs.exclude(pk=previous_current_id)
        stale_in_auction_qs.update(status=Player.Status.AVAILABLE)
        player.status = Player.Status.IN_AUCTION
        player.save(update_fields=["status"])
        auction.current_player = player
        auction.status = Auction.Status.LIVE
        auction.sold_animation_state = False
        auction.save(update_fields=["current_player", "status", "sold_animation_state"])
        AuctionLog.objects.create(auction=auction, actor=request.user, action="auction.current_player", message=f"{player.full_name} is now live.")
        bump_live_revision(auction)
        return live_response(serialize_live_state(auction))

    @action(detail=True, methods=["post"], url_path="manual-bid")
    def manual_bid(self, request, auction_id=None):
        auction = self.get_object()
        require_auction_staff(request.user)
        return self._create_bid(request, auction, Bid.BidType.MANUAL)

    @action(detail=True, methods=["post"], url_path="team-owner-bid")
    def team_owner_bid(self, request, auction_id=None):
        if not is_team_owner(request.user):
            raise PermissionDenied("Only team owners can bid from the owner screen.")
        auction = self.get_object()
        auction_settings, _ = AuctionSettings.objects.get_or_create(auction=auction)
        if not auction_settings.enable_owner_bidding:
            raise PermissionDenied("Owner bidding is disabled for this project.")
        return self._create_bid(request, auction, Bid.BidType.TEAM_OWNER)

    def _create_bid(self, request, auction: Auction, bid_type: str):
        player = auction.current_player
        if not player:
            raise ValidationError({"player": "Set a current player before bidding."})
        team_value = request.data.get("team") or request.data.get("team_id")
        if is_team_owner(request.user):
            profile = request.user.role_profile
            team = profile.team
        else:
            team = auction.teams.filter(team_id=team_value).first() or auction.teams.filter(pk=team_value).first()
        if not team or team.auction_id != auction.pk:
            raise ValidationError({"team": "Team not found in this auction."})
        if team.status != Team.Status.ACTIVE:
            raise ValidationError(
                {
                    "team": (
                        f"{team.short_name or team.name} is "
                        f"{team.get_status_display().lower()} and cannot place bids."
                    )
                }
            )
        validate_team_player_limits(team, player)
        raw_amount = request.data.get("bid_amount") or request.data.get("amount") or "0"
        try:
            amount = Decimal(str(raw_amount).replace(",", "").strip())
        except (InvalidOperation, ValueError) as exc:
            raise ValidationError({"bid_amount": "Bid amount must be a valid number."}) from exc
        if amount <= 0:
            raise ValidationError({"bid_amount": "Bid amount must be greater than zero."})
        budget = available_bid_budget(team)
        if amount > budget:
            raise ValidationError(
                {"bid_amount": f"Bid cannot exceed available budget after reserve ({budget})."}
            )
        current = highest_active_bid(auction, player)
        team_current = highest_active_bid_for_team(auction, player, team)
        if team_current:
            minimum = (current.bid_amount if current else player.base_price) + auction.bid_increment
        else:
            minimum = player.base_price
        if amount < minimum:
            raise ValidationError({"bid_amount": f"Bid must be at least {minimum}."})
        bid = Bid.objects.create(
            auction=auction,
            player=player,
            team=team,
            bid_amount=amount,
            bid_type=bid_type,
            bid_status=Bid.Status.PENDING,
        )
        AuctionLog.objects.create(
            auction=auction,
            actor=request.user,
            action="bid.pending",
            message=f"{team.short_name} bid {amount} for {player.full_name}.",
        )
        bump_live_revision(auction)
        return live_response(BidSerializer(bid).data, status.HTTP_201_CREATED)

    @action(detail=True, methods=["get"], url_path="pending-bids")
    def pending_bids(self, request, auction_id=None):
        require_auction_staff(request.user)
        auction = self.get_object()
        bids = auction.bids.select_related("player", "player__category", "team").filter(bid_status=Bid.Status.PENDING)
        return live_response(BidSerializer(bids, many=True).data)

    @transaction.atomic
    @action(detail=True, methods=["post"], url_path=r"bids/(?P<bid_pk>[^/.]+)/approve")
    def approve_bid(self, request, auction_id=None, bid_pk=None):
        require_auction_staff(request.user)
        auction = self.get_object()
        bid = auction.bids.select_related("player", "team").get(pk=bid_pk)
        if bid.bid_status != Bid.Status.PENDING:
            raise ValidationError({"bid": "Only pending bids can be approved."})
        current_highest = highest_active_bid(auction, bid.player)
        if not current_highest or current_highest.pk != bid.pk:
            raise ValidationError({"bid": "Only the current highest bid can be approved."})
        validate_team_player_limits(bid.team, bid.player)
        budget = available_bid_budget(bid.team)
        if bid.bid_amount > budget:
            raise ValidationError(
                {"bid_amount": f"Winning team must keep reserved purse. Available bid budget is {budget}."}
            )
        bid.approve(request.user)
        AuctionLog.objects.create(auction=auction, actor=request.user, action="bid.approved", message=f"Approved {bid.team.short_name} bid.")
        if bid.player.status not in {Player.Status.SOLD, Player.Status.UNSOLD}:
            sell_current_player_to_team(
                auction,
                bid.player,
                bid.team,
                bid.bid_amount,
                actor=request.user,
                winning_bid=bid,
            )
        else:
            bump_live_revision(auction)
        return live_response(serialize_live_state(auction))

    @action(detail=True, methods=["post"], url_path=r"bids/(?P<bid_pk>[^/.]+)/reject")
    def reject_bid(self, request, auction_id=None, bid_pk=None):
        require_auction_staff(request.user)
        auction = self.get_object()
        bid = auction.bids.get(pk=bid_pk)
        bid.reject(request.user)
        AuctionLog.objects.create(auction=auction, actor=request.user, action="bid.rejected", message=f"Rejected {bid.team.short_name} bid.")
        bump_live_revision(auction)
        return live_response(BidSerializer(bid).data)

    @transaction.atomic
    @action(detail=True, methods=["post"], url_path="sell-player")
    def sell_player(self, request, auction_id=None):
        require_auction_staff(request.user)
        auction = self.get_object()
        player = auction.current_player
        if not player:
            raise ValidationError({"player": "No current player selected."})
        winning_bid = highest_approved_bid(auction, player)
        team_value = request.data.get("team") or request.data.get("team_id")
        team = None
        if team_value:
            team = auction.teams.filter(team_id=team_value).first() or auction.teams.filter(pk=team_value).first()
        if not team and winning_bid:
            team = winning_bid.team
        if not team:
            raise ValidationError({"team": "Approve a bid or choose a winning team."})
        sold_price = Decimal(str(request.data.get("sold_price") or request.data.get("amount") or (winning_bid.bid_amount if winning_bid else player.base_price)))
        sell_current_player_to_team(
            auction,
            player,
            team,
            sold_price,
            actor=request.user,
            winning_bid=winning_bid,
        )
        return live_response(serialize_live_state(auction))

    @action(detail=True, methods=["post"], url_path="mark-unsold")
    def mark_unsold(self, request, auction_id=None):
        require_auction_staff(request.user)
        auction = self.get_object()
        player = auction.current_player
        if not player:
            raise ValidationError({"player": "No current player selected."})
        player.status = Player.Status.UNSOLD
        player.sold_team = None
        player.sold_price = None
        player.save(update_fields=["status", "sold_team", "sold_price"])
        auction.current_player = None
        auction.sold_animation_state = False
        auction.save(update_fields=["current_player", "sold_animation_state"])
        Bid.objects.filter(auction=auction, player=player, bid_status=Bid.Status.PENDING).update(
            bid_status=Bid.Status.REJECTED,
            approved_by_admin=request.user,
            approved_at=timezone.now(),
        )
        AuctionLog.objects.create(auction=auction, actor=request.user, action="player.unsold", message=f"{player.full_name} marked unsold.")
        bump_live_revision(auction)
        return live_response(serialize_live_state(auction))

    @action(detail=True, methods=["post"], url_path="next-player")
    def next_player(self, request, auction_id=None):
        require_auction_staff(request.user)
        auction = self.get_object()
        advance_to_random_player(auction, request.user)
        bump_live_revision(auction)
        return live_response(serialize_live_state(auction))

    @transaction.atomic
    @action(detail=True, methods=["post"], url_path="move-current-player-category")
    def move_current_player_category(self, request, auction_id=None):
        require_auction_staff(request.user)
        auction = self.get_object()
        player = auction.current_player
        if not player:
            raise ValidationError({"player": "No current player selected."})

        source_category = player.category
        target_category = resolve_auction_category(
            auction,
            request.data.get("to_category") or request.data.get("target_category"),
            "to_category",
        )
        if player.category_id == target_category.pk:
            raise ValidationError({"to_category": "Choose a different category to move this player into."})

        player.category = target_category
        player.base_price = target_category.base_value
        player.save(update_fields=["category", "base_price"])
        auction.current_player = player
        AuctionLog.objects.create(
            auction=auction,
            actor=request.user,
            action="player.category_moved",
            message=(
                f"Moved {player.full_name} from "
                f"{source_category.name if source_category else 'No category'} to {target_category.name}."
            ),
            metadata={
                "player": player.pk,
                "from_category": source_category.pk if source_category else None,
                "to_category": target_category.pk,
            },
        )
        bump_live_revision(auction)
        return live_response(
            {
                "from_category": source_category.pk if source_category else None,
                "to_category": target_category.pk,
                "state": serialize_live_state(auction),
            }
        )

    @transaction.atomic
    @action(detail=True, methods=["post"], url_path="move-remaining-players")
    def move_remaining_players(self, request, auction_id=None):
        require_auction_staff(request.user)
        auction = self.get_object()
        source_category = resolve_auction_category(
            auction,
            request.data.get("from_category") or request.data.get("source_category"),
            "from_category",
        )
        target_category = resolve_auction_category(
            auction,
            request.data.get("to_category") or request.data.get("target_category"),
            "to_category",
        )
        if source_category.pk == target_category.pk:
            raise ValidationError({"to_category": "Choose a different category to move players into."})

        players_qs = auction.players.filter(
            category=source_category,
            status=Player.Status.AVAILABLE,
        )
        moved_count = players_qs.update(
            category=target_category,
            base_price=target_category.base_value,
        )
        if moved_count:
            AuctionLog.objects.create(
                auction=auction,
                actor=request.user,
                action="players.category_moved",
                message=(
                    f"Moved {moved_count} remaining players from "
                    f"{source_category.name} to {target_category.name}."
                ),
                metadata={
                    "from_category": source_category.pk,
                    "to_category": target_category.pk,
                    "moved_count": moved_count,
                },
            )
            bump_live_revision(auction)

        return live_response(
            {
                "moved_count": moved_count,
                "state": serialize_live_state(auction),
            }
        )

    @action(detail=True, methods=["get"], url_path="team-roster")
    def team_roster(self, request, auction_id=None):
        auction = self.get_object()
        team = self._roster_team_for_request(request, auction)
        players = auction.players.filter(sold_team=team).select_related("category", "sold_team").order_by("full_name", "id")
        return Response({"team": TeamSerializer(team).data, "players": PlayerSerializer(players, many=True).data})

    @action(detail=True, methods=["get"], url_path="team-rosters")
    def team_rosters(self, request, auction_id=None):
        require_auction_staff(request.user)
        auction = self.get_object()
        teams = list(live_team_queryset(auction).order_by("name"))
        team_by_pk = {team.pk: team for team in teams}
        sold_records = (
            auction.sold_players.filter(team_id__in=team_by_pk)
            .select_related("player", "player__category", "team")
            .order_by("team__name", "player__full_name", "player__id")
        )
        players_by_team: dict[int, list[dict]] = defaultdict(list)
        for sold in sold_records:
            player_data = live_player_data(sold.player)
            if player_data:
                player_data["sold_team"] = sold.team_id
                player_data["sold_team_name"] = sold.team.name
                player_data["sold_price"] = decimal_text(sold.sold_price)
                players_by_team[sold.team_id].append(player_data)
        return Response(
            {
                "auction": live_auction_data(auction, build_results(auction, teams=teams), len(teams)),
                "rosters": [
                    {
                        "team": live_team_data(team),
                        "players": players_by_team.get(team.pk, []),
                    }
                    for team in teams
                ],
            }
        )

    @action(detail=True, methods=["get"], url_path="team-roster-pdf")
    def team_roster_pdf(self, request, auction_id=None):
        auction = self.get_object()
        team = self._roster_team_for_request(request, auction)
        pdf_data = build_team_roster_pdf(auction, team, request=request)
        filename = f"{team_roster_pdf_filename(auction, team)}.pdf"
        response = HttpResponse(pdf_data, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Content-Length"] = str(len(pdf_data))
        return response

    def _roster_team_for_request(self, request, auction: Auction) -> Team:
        if is_team_owner(request.user):
            team = request.user.role_profile.team
            if not team or team.auction_id != auction.pk:
                raise PermissionDenied("This owner is not assigned to a team in this auction.")
        else:
            team_value = request.query_params.get("team") or request.query_params.get("team_id")
            team = auction.teams.filter(team_id=team_value).first() or auction.teams.filter(pk=team_value).first()
        if not team:
            raise ValidationError({"team": "Team is required."})
        return team

    @action(detail=True, methods=["get"], url_path="sold-players")
    def sold_players(self, request, auction_id=None):
        auction = self.get_object()
        return live_response(SoldPlayerSerializer(auction.sold_players.select_related("player", "player__category", "team"), many=True).data)

    @action(detail=True, methods=["get"], url_path="results")
    def results(self, request, auction_id=None):
        return live_response(build_results(self.get_object()))

    @action(detail=True, methods=["get"], url_path="public-live")
    def public_live(self, request, auction_id=None):
        return live_response(serialize_live_state(self.get_object()))

    @action(detail=False, methods=["get"], url_path="public-active")
    def public_active(self, request):
        auction = public_active_auction()
        if not auction:
            return Response({"detail": "No active auction is available."}, status=status.HTTP_404_NOT_FOUND)
        return live_response(serialize_live_state(auction))

    @action(detail=True, methods=["get"], url_path="projector")
    def projector(self, request, auction_id=None):
        state = serialize_live_state(self.get_object())
        state["mode"] = "projector"
        return live_response(state)


class CategoryViewSet(ScopedModelViewSet):
    queryset = Category.objects.select_related("auction")
    serializer_class = CategorySerializer

    def perform_create(self, serializer):
        super().perform_create(serializer)
        category = serializer.instance
        if category.maximum_players > 0:
            TeamCategoryLimit.objects.bulk_create(
                [
                    TeamCategoryLimit(
                        team=team,
                        category=category,
                        maximum_players=category.maximum_players,
                    )
                    for team in category.auction.teams.all()
                ],
                ignore_conflicts=True,
            )

    def perform_update(self, serializer):
        super().perform_update(serializer)
        category = serializer.instance
        if "maximum_players" in serializer.validated_data:
            for team in category.auction.teams.all():
                TeamCategoryLimit.objects.update_or_create(
                    team=team,
                    category=category,
                    defaults={"maximum_players": category.maximum_players},
                )


class PlayerViewSet(ScopedModelViewSet):
    queryset = Player.objects.select_related("auction", "category", "sold_team")
    serializer_class = PlayerSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        player_status = self.request.query_params.get("status")
        valid_statuses = {choice[0] for choice in Player.Status.choices}
        if player_status in valid_statuses:
            qs = qs.filter(status=player_status)
        return qs

    def perform_create(self, serializer):
        super().perform_create(serializer)
        if "base_price" not in serializer.validated_data:
            self.sync_player_base_price(serializer.instance)

    def perform_update(self, serializer):
        old_category_id = serializer.instance.category_id
        super().perform_update(serializer)
        category_changed = "category" in serializer.validated_data and serializer.instance.category_id != old_category_id
        if category_changed and "base_price" not in serializer.validated_data:
            self.sync_player_base_price(serializer.instance)

    def sync_player_base_price(self, player):
        base_price = player.category.base_value if player.category else Decimal("0")
        if player.base_price != base_price:
            player.base_price = base_price
            player.save(update_fields=["base_price"])

    @transaction.atomic
    @action(detail=False, methods=["delete"], url_path="bulk-delete")
    def bulk_delete(self, request):
        require_auction_staff(request.user)
        auction_value = request.data.get("auction") or request.query_params.get("auction")
        auction = None
        if auction_value:
            auction = Auction.objects.filter(auction_id=auction_value).first()
            if not auction and str(auction_value).isdigit():
                auction = Auction.objects.filter(pk=int(auction_value)).first()
        if not auction:
            auction = scoped_auction_for_user(request.user)
        if not auction:
            raise ValidationError({"auction": "Choose an auction before deleting players."})
        if not user_can_access_auction(request.user, auction):
            raise PermissionDenied("You cannot delete players for this auction.")

        player_count = auction.players.count()
        auction.current_player = None
        auction.sold_animation_state = False
        auction.save(update_fields=["current_player", "sold_animation_state"])
        auction.players.all().delete()
        auction.teams.update(players_bought=0, purse_amount=auction.purse_amount, remaining_purse=auction.purse_amount)
        bump_live_revision(auction)
        AuctionLog.objects.create(
            auction=auction,
            actor=request.user,
            action="players.bulk_delete",
            message=f"Deleted {player_count} players.",
            metadata={"deleted_count": player_count},
        )
        return Response({"deleted_count": player_count})

    @action(detail=False, methods=["post"], url_path="import-excel")
    def import_excel(self, request):
        require_auction_staff(request.user)
        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            raise ValidationError({"file": "Choose an Excel or CSV file to upload."})

        suffix = Path(uploaded_file.name).suffix.lower()
        if suffix not in {".xlsx", ".xlsm", ".csv"}:
            raise ValidationError({"file": "Upload an .xlsx, .xlsm, or .csv file."})

        auction_value = request.data.get("auction") or request.query_params.get("auction")
        auction = None
        if auction_value:
            auction = Auction.objects.filter(pk=auction_value).first() or Auction.objects.filter(auction_id=auction_value).first()
        if not auction:
            auction = scoped_auction_for_user(request.user)
        if not auction:
            raise ValidationError({"auction": "Choose an auction before importing players."})
        if not user_can_access_auction(request.user, auction):
            raise PermissionDenied("You cannot import players for this auction.")

        workbook = None
        try:
            if suffix == ".csv":
                raw_file = uploaded_file.read()
                try:
                    csv_text = raw_file.decode("utf-8-sig")
                except UnicodeDecodeError:
                    csv_text = raw_file.decode("latin-1")
                rows = csv.reader(csv_text.splitlines())
            else:
                try:
                    workbook = load_workbook(uploaded_file, read_only=False, data_only=True)
                except Exception as exc:
                    raise ValidationError({"file": "The Excel file could not be read."}) from exc
                sheet = workbook.active
                rows = sheet.iter_rows()

            header = next(rows, None)
            if not header:
                raise ValidationError({"file": "The import file is empty."})

            columns = {}
            for index, heading in enumerate(header):
                key = canonical_import_key(heading)
                if key and key not in columns:
                    columns[key] = index

            if "first_name" not in columns and "full_name" not in columns:
                raise ValidationError({"file": "Add a Name or First Name column to the Excel sheet."})

            def row_value(row, key):
                index = columns.get(key)
                if index is None or index >= len(row):
                    return ""
                return row[index]

            created_players = []
            created_category_ids = set()
            errors = []
            for row_number, row in enumerate(rows, start=2):
                if not any(clean_import_cell(cell) for cell in row):
                    continue
                try:
                    full_name = clean_import_cell(row_value(row, "full_name"))
                    first_name = clean_import_cell(row_value(row, "first_name"))
                    last_name = clean_import_cell(row_value(row, "last_name"))
                    if full_name and not first_name:
                        parts = full_name.split()
                        first_name = parts[0]
                        last_name = " ".join(parts[1:])
                    if not first_name:
                        raise ValueError("Name or First Name is required.")
                    base_price_value = row_value(row, "base_price")
                    base_price_text = clean_import_cell(base_price_value)
                    imported_base_price = parse_import_decimal(base_price_value, "0") if base_price_text else Decimal("0")
                    category, category_created = resolve_import_category(auction, row_value(row, "category"), imported_base_price)
                    if category_created and category:
                        created_category_ids.add(category.pk)
                    player_base_price = imported_base_price if base_price_text else (category.base_value if category else Decimal("0"))

                    player = Player.objects.create(
                        auction=auction,
                        first_name=first_name[:80],
                        last_name=last_name[:80],
                        category=category,
                        image_url=clean_import_url(row_value(row, "image_url")),
                        role=resolve_import_role(row_value(row, "role")),
                        country=clean_import_cell(row_value(row, "country"))[:80],
                        age=max(parse_import_int(row_value(row, "age"), 0), 0),
                        base_price=player_base_price,
                        queue_order=max(parse_import_int(row_value(row, "queue_order"), 0), 0),
                    )
                    created_players.append(player)
                except ValueError as exc:
                    errors.append({"row": row_number, "message": str(exc)})

            AuctionLog.objects.create(
                auction=auction,
                actor=request.user,
                action="players.import",
                message=f"Imported {len(created_players)} players from import file.",
                metadata={
                    "file": uploaded_file.name,
                    "created_count": len(created_players),
                    "created_category_count": len(created_category_ids),
                    "error_count": len(errors),
                },
            )
            return Response(
                {
                    "created_count": len(created_players),
                    "created_category_count": len(created_category_ids),
                    "error_count": len(errors),
                    "errors": errors[:50],
                    "players": PlayerSerializer(created_players[:25], many=True).data,
                },
                status=status.HTTP_201_CREATED,
            )
        finally:
            if workbook:
                workbook.close()


class TeamViewSet(ScopedModelViewSet):
    queryset = Team.objects.select_related("auction", "owner_user").prefetch_related(
        Prefetch("category_limits", queryset=TeamCategoryLimit.objects.select_related("category")),
        Prefetch("players", queryset=Player.objects.only("id", "sold_team_id", "category_id")),
    )
    serializer_class = TeamSerializer

    @transaction.atomic
    @action(detail=True, methods=["post"], url_path="category-limits")
    def category_limits(self, request, pk=None):
        require_auction_staff(request.user)
        team = self.get_object()
        raw_limits = request.data.get("limits", [])
        if not isinstance(raw_limits, list):
            raise ValidationError({"limits": "Send limits as a list."})

        seen_categories = set()
        total_maximum_players = 0
        for item in raw_limits:
            if not isinstance(item, dict):
                raise ValidationError({"limits": "Every limit must be an object."})
            category_value = item.get("category")
            maximum_value = item.get("maximum_players", 0)
            try:
                maximum_players = max(int(maximum_value or 0), 0)
            except (TypeError, ValueError) as exc:
                raise ValidationError({"maximum_players": "Max players must be a number."}) from exc
            category = team.auction.categories.filter(category_id=category_value).first()
            if not category and str(category_value).isdigit():
                category = team.auction.categories.filter(pk=int(category_value)).first()
            if not category:
                raise ValidationError({"category": "Category not found for this auction."})
            if category.pk in seen_categories:
                raise ValidationError({"category": f"{category.name} was included more than once."})
            bought_count = team.players.filter(category=category).count()
            if maximum_players and maximum_players < bought_count:
                raise ValidationError(
                    {
                        "maximum_players": (
                            f"{team.short_name} already has {bought_count} players in {category.name}. "
                            "Max players for this category cannot be lower than that."
                        )
                    }
                )
            seen_categories.add(category.pk)
            total_maximum_players += maximum_players
            if maximum_players:
                TeamCategoryLimit.objects.update_or_create(
                    team=team,
                    category=category,
                    defaults={"maximum_players": maximum_players},
                )
            else:
                TeamCategoryLimit.objects.filter(team=team, category=category).delete()

        if team.maximum_players and total_maximum_players > team.maximum_players:
            raise ValidationError(
                {
                    "limits": (
                        f"Category limits total {total_maximum_players}, but {team.short_name}'s squad maximum "
                        f"is {team.maximum_players}."
                    )
                }
            )

        TeamCategoryLimit.objects.filter(team=team).exclude(category_id__in=seen_categories).delete()
        AuctionLog.objects.create(
            auction=team.auction,
            actor=request.user,
            action="team.category_limits",
            message=f"Updated category limits for {team.short_name}.",
        )
        return Response(TeamSerializer(team).data)

    def perform_destroy(self, instance):
        require_auction_staff(self.request.user)
        if not user_can_access_auction(self.request.user, instance.auction):
            raise PermissionDenied("You cannot delete this team.")
        owner_user = instance.owner_user
        if owner_user and hasattr(owner_user, "role_profile"):
            owner_user.role_profile.delete()
        instance.delete()


class TeamOwnerViewSet(ScopedModelViewSet):
    queryset = TeamOwner.objects.select_related("auction", "team", "user")
    serializer_class = TeamOwnerSerializer


class SponsorViewSet(ScopedModelViewSet):
    queryset = Sponsor.objects.select_related("auction")
    serializer_class = SponsorSerializer


class BidViewSet(mixins.ListModelMixin, mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    queryset = Bid.objects.select_related("auction", "player", "team")
    serializer_class = BidSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        if is_super_admin(self.request.user):
            return qs
        auction = scoped_auction_for_user(self.request.user)
        if not auction:
            return qs.none()
        if is_team_owner(self.request.user):
            return qs.filter(auction=auction, team=self.request.user.role_profile.team)
        return qs.filter(auction=auction)


class SoldPlayerViewSet(mixins.ListModelMixin, mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    queryset = SoldPlayer.objects.select_related("auction", "player", "team")
    serializer_class = SoldPlayerSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        if is_super_admin(self.request.user):
            return qs
        auction = scoped_auction_for_user(self.request.user)
        return qs.filter(auction=auction) if auction else qs.none()


class AuctionLogViewSet(mixins.ListModelMixin, viewsets.GenericViewSet):
    queryset = AuctionLog.objects.select_related("auction", "actor")
    serializer_class = AuctionLogSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        if is_super_admin(self.request.user):
            return qs
        auction = scoped_auction_for_user(self.request.user)
        return qs.filter(auction=auction) if auction else qs.none()
